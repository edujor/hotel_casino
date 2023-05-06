import tkinter
from tkinter import messagebox
import datetime
from openpyxl import load_workbook,Workbook
import win32com.client
import datetime
import locale
from pathlib import Path
from openpyxl.styles import Font, Alignment

def main_excel(filename,year,month):
    file_path = Path(filename)
    # Verifica si el archivo ya existe
    if file_path.exists():
        # Si el archivo existe, carga el libro de trabajo
        workbook = load_workbook(filename)
    else:
        print('Creando el excel ...')
        # Si el archivo no existe, crea un nuevo libro de trabajo
        workbook = Workbook()

        # worksheet = workbook.active
        crear_excel(workbook,year,month)
        workbook.save(filename)
        sheet_names = workbook.sheetnames
        print(sheet_names)
    return workbook

def crear_excel(workbook,year,month):
    lista_dias_inicial = conseguir_fechas(year,month)
    for dia in lista_dias_inicial:
        # dia_str = dia
        hoja = workbook.create_sheet(str(dia))
        # Creando la 1° Fila
        cell = hoja['B1']
        cell.value = 'HAB'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['C1']
        cell.value = 'TIPO DE HAB'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['D1']
        cell.value = 'ESTADO'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['E1']
        cell.value = 'ID'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['F1']
        cell.value = 'APELLIDOS'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['G1']
        cell.value = 'NOMBRES'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['H1']
        cell.value = 'N° BOLETA'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['I1']
        cell.value = 'TELEFONO'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['J1']
        cell.value = 'F. INGRESO'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['K1']
        cell.value = 'F. SALIDA'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['L1']
        cell.value = 'N° HUESPEDES'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['M1']
        cell.value = 'F. DE PAGO'
        aplicar_formato(cell,'negrita','centrado')
        cell = hoja['N1']
        cell.value = 'OBSERVACIÓN'
        aplicar_formato(cell,'negrita','centrado')

        # Creando el resto de filas
        llenando_datos_iniciales(hoja)

def llenando_datos_iniciales(sheet):
    for i in range(2, 103):
    # Insertar número en columna 1
        sheet.cell(row=i, column=1, value=i-1)
    
        # Determinar valor para columna 2
        if i in [2,34,35,68,69,76,95]:
            valor = "SIMPLE"
        elif i in [10,13,20,21,22,25,41,42,56,61,62]:
            valor = "MATRIMONIAL"
        elif i in [7,8,27,28]:
            valor = "SUITE"
        elif i in [11,12,14,15,23,24,26,43,44,45,46,47,48,49,54,55,57,58,59,60,77,78,79,80,81,82,83,88,89,90,91,92,93,94]:
            valor = "TRIPLE"
        else:
            valor = "DOBLE"
        # Insertar valor en columna 2
        sheet.cell(row=i, column=2, value=valor)


def aplicar_formato(cell, fuente, alineacion):
    # Crear objeto de fuente
    if fuente == 'negrita':
        fuente_obj = Font(bold=True)
    elif fuente == 'cursiva':
        fuente_obj = Font(italic=True)
    elif fuente == 'subrayado':
        fuente_obj = Font(underline='single')
    else:
        fuente_obj = Font()

    # Crear objeto de alineación
    if alineacion == 'centrado':
        alineacion_obj = Alignment(horizontal='center')
    elif alineacion == 'derecha':
        alineacion_obj = Alignment(horizontal='right')
    elif alineacion == 'izquierda':
        alineacion_obj = Alignment(horizontal='left')
    else:
        alineacion_obj = Alignment()

    # Asignar fuente y alineación a la celda
    cell.font = fuente_obj
    cell.alignment = alineacion_obj
def conseguir_fechas(year,month):

    # Establecer la configuración regional en español
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

    # Calcula el primer día del mes
    first_day = datetime.date(year, month, 1)

    # Obtiene el nombre abreviado del día de la semana en español
    def get_weekday_abbr(date):
        return date.strftime("%A").upper()[:3].replace('SÁB', 'SAB').replace('VIE', 'VIE').replace('JUE', 'JUE').replace('MIÉ', 'MIE').replace('MAR', 'MAR').replace('LUN', 'LUN').replace('DOM', 'DOM')

    # Genera una lista de fechas y días de la semana en el formato requerido con nombres de días en español y formato corregido
    dates = [(get_weekday_abbr(first_day + datetime.timedelta(days=i)), str(first_day.day + i)) for i in range(0, 31) if first_day.month == (first_day + datetime.timedelta(days=i)).month]

    # Imprime la lista de fechas y días de la semana
    return dates

def leer_excel(filename,tk,window):
    active_sheet_name = ""
    # Carga el archivo de Excel
    workbook = load_workbook(filename)
    # Obtén una lista de las hojas del libro de trabajo
    sheet_names = workbook.sheetnames
    # Crea una pestaña en la ventana de tkinter para cada hoja del libro de trabajo
    for sheet_name in sheet_names:
        print('Pestaña actual → ',sheet_name)
        # Crea una pestaña para la hoja actual
        tab = tk.Frame(window)
        tab.pack(fill="both", expand=True)
        
        # Asigna el nombre de la hoja activa a la variable "active_sheet_name"
        active_sheet_name = sheet_name    
        
        # Obtén una referencia a la hoja actual
        sheet = workbook[sheet_name]
        
        # Obtén los datos de la hoja actual
        data = []
        print('obteniendo los datos de cada hoja')
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
    
        # Crea una tabla en la pestaña para mostrar los datos
        for i, row in enumerate(data):
            row_labels = []  # Lista de etiquetas para esta fila
            for j in range(len(row)):
                label_text = str(row[j]) if row[j] is not None else ""
                label = tk.Label(tab, text=label_text)
                row_labels.append(label)
    # Agrega todas las etiquetas de esta fila a la tabla
            for j, label in enumerate(row_labels):
                label.grid(row=i, column=j)
    workbook.close()

def guardar_datos(tipo_habitacion_combobox,estado_combobox,num_doc_entrada,nombre_entrada,apellido_entrada,num_boleta_entrada,telefono_entrada,fecha_ingreso_entrada,fecha_salida_entrada,observacion_entrada,cond_indefinido,fecha_salida_indefinida):
    # Obtener los valores ingresados
    tipo_habitacion_value = tipo_habitacion_combobox.get()
    estado_value = estado_combobox.get()
    num_doc_value = num_doc_entrada.get()
    nombre_value = nombre_entrada.get()
    apellido_value = apellido_entrada.get()
    num_boleta_value = num_boleta_entrada.get()
    telefono_value = telefono_entrada.get()
    fecha_ingreso_value = fecha_ingreso_entrada.get()
    fecha_salida_value = fecha_salida_entrada.get()
    observacion_value = observacion_entrada.get()

    mensaje = f"Tipo de Habitación: {tipo_habitacion_value}\nEstado: {estado_value}\nN° Documento: {num_doc_value}\nNombres: {nombre_value}\nApellidos: {apellido_value}\nN° Boleta: {num_boleta_value}\nTeléfono: {telefono_value}\nFecha de Ingreso: {fecha_ingreso_value}\nFecha de Salida: {fecha_salida_value}\nObservación: {observacion_value}"
    print(mensaje)
    tkinter.messagebox.showinfo("Datos ingresados", mensaje)
    
    llenar_excel(tipo_habitacion_value,estado_value,num_doc_value,nombre_value,apellido_value,num_boleta_value,telefono_value,fecha_ingreso_value,fecha_salida_value,observacion_value,cond_indefinido)
    # Limpiar los campos
    limpiar_campos(tipo_habitacion_combobox,estado_combobox,num_doc_entrada,nombre_entrada,apellido_entrada,num_boleta_entrada,telefono_entrada,fecha_ingreso_entrada,fecha_salida_entrada,observacion_entrada,fecha_salida_indefinida)

def limpiar_campos(tipo_habitacion_combobox,estado_combobox,num_doc_entrada,nombre_entrada,apellido_entrada,num_boleta_entrada,telefono_entrada,fecha_ingreso_entrada,fecha_salida_entrada,observacion_entrada,fecha_salida_indefinida):
    tipo_habitacion_combobox.set("")
    estado_combobox.set("")
    num_doc_entrada.delete(0, tkinter.END)
    nombre_entrada.delete(0, tkinter.END)
    apellido_entrada.delete(0, tkinter.END)
    num_boleta_entrada.delete(0, tkinter.END)
    telefono_entrada.delete(0, tkinter.END)
    hoy = datetime.date.today().strftime('%d/%m/%Y')
    fecha_ingreso_entrada.set_date(hoy)
    fecha_salida_entrada.set_date(hoy)
    observacion_entrada.delete(0, tkinter.END)
    # Desmarcar la opción de "fecha indefinida"
    fecha_salida_indefinida.deselect()
    #fecha_ingreso_entrada.set_date("")
    #fecha_salida_entrada.set_date("")

def crear_widgets(user_info_frame,DateEntry):

    fecha_salida_indefinida_var = tkinter.BooleanVar()

    # Crear widget de la fecha de ingreso
    fecha_ingreso = tkinter.Label(user_info_frame, text="Fecha Ingreso")
    fecha_ingreso.grid(row=4, column=1)

    fecha_ingreso_entrada = DateEntry(user_info_frame, date_pattern='dd/mm/yyyy')
    fecha_ingreso_entrada.grid(row=5, column=1)

    # Crear widget de la fecha de salida
    fecha_salida = tkinter.Label(user_info_frame, text="Fecha Salida")
    fecha_salida.grid(row=4, column=2)

    fecha_salida_entrada = DateEntry(user_info_frame, date_pattern='dd/mm/yyyy')
    fecha_salida_entrada.grid(row=5, column=2)

    # Crear widget para la opción de fecha de salida indefinida
    # fecha_salida_indefinida = tkinter.Checkbutton(user_info_frame, text="Fecha de salida indefinida")
    cond_indefinido = tkinter.BooleanVar()
    fecha_salida_indefinida = tkinter.Checkbutton(user_info_frame, text="Fecha de salida indefinida", variable=cond_indefinido)
    fecha_salida_indefinida.grid(row=6, column=3)
    
    # Definir función para actualizar fecha de salida mínima
    def actualizar_fecha_salida_minimo(evento):
        try:
            fecha_ingreso = evento.widget.get_date()
            if fecha_salida_indefinida_var.get():
                fecha_salida_entrada.delete(0, tkinter.END)
            else:
                fecha_salida_entrada.config(
                    mindate=fecha_ingreso,
                    year=fecha_ingreso.year,
                    month=fecha_ingreso.month,
                    day=fecha_ingreso.day
                    
                )
                fecha_salida_entrada.config(validate="key") 
            fecha_salida_entrada.set_date(fecha_salida_entrada.get_date())

        except AttributeError:
            pass
    # Vincular función con el evento de selección de fecha de ingreso
    fecha_ingreso_entrada.bind("<<DateEntrySelected>>", actualizar_fecha_salida_minimo)
    # Devolver los widgets
    
    return fecha_ingreso_entrada, fecha_salida_entrada ,cond_indefinido,fecha_salida_indefinida

def llenar_excel(tipo_habitacion_value,estado_value,num_doc_value,nombre_value,apellido_value,num_boleta_value,telefono_value,fecha_ingreso_value,fecha_salida_value,observacion_value,cond_indefinido):
    # Obtiene la instancia de Excel que está utilizando el archivo
    #excel = win32com.client.GetActiveObject('Excel.Application')

    excel = win32com.client.Dispatch('Excel.Application')

    # Abre el archivo de Excel en modo de lectura y escritura
    workbook = excel.Workbooks.Open('hotel.xlsx')

    # Selecciona la hoja de trabajo correspondiente
    if 'Registro de clientes' in [sheet.Name for sheet in workbook.Sheets]:
        worksheet = workbook.Sheets('Registro de clientes')
    else:
        worksheet = workbook.Sheets.Add()
        worksheet.Name = 'Registro de clientes'

    # Obtiene la última fila con datos y agrega la nueva fila debajo
    last_row = worksheet.Cells.SpecialCells(11).Row
    new_row = last_row + 1

    # Agrega los datos a la nueva fila
    worksheet.Cells(new_row, 1).Value = tipo_habitacion_value
    worksheet.Cells(new_row, 2).Value = estado_value
    worksheet.Cells(new_row, 3).Value = num_doc_value
    worksheet.Cells(new_row, 4).Value = nombre_value
    worksheet.Cells(new_row, 5).Value = apellido_value
    worksheet.Cells(new_row, 6).Value = num_boleta_value
    worksheet.Cells(new_row, 7).Value = telefono_value
    worksheet.Cells(new_row, 8).Value = fecha_ingreso_value

    if not cond_indefinido.get():
        worksheet.Cells(new_row, 9).Value = fecha_salida_value
    else:
        worksheet.Cells(new_row, 9).Value = "Indefinido"

    worksheet.Cells(new_row, 10).Value = observacion_value

    # Guarda los cambios en el archivo de Excel
    workbook.Save()
    
    '''# Carga el archivo de Excel
    workbook = load_workbook(filename='hotel.xlsx')

    # Crea una nueva hoja en el archivo de Excel
    if 'Registro de clientes' in workbook.sheetnames:
        sheet = workbook['Registro de clientes']
    else:
        sheet = workbook.create_sheet('Registro de clientes')
    
    row = sheet.max_row + 1
    sheet[f'A{row}'] = tipo_habitacion_value
    sheet[f'B{row}'] = estado_value
    sheet[f'C{row}'] = num_doc_value
    sheet[f'D{row}'] = nombre_value
    sheet[f'E{row}'] = apellido_value
    sheet[f'F{row}'] = num_boleta_value
    sheet[f'G{row}'] = telefono_value
    sheet[f'H{row}'] = fecha_ingreso_value
    
    if not cond_indefinido.get():
        sheet[f'I{row}'] = fecha_salida_value
    else:
        sheet[f'I{row}'] = "Indefinido"
    sheet[f'J{row}'] = observacion_value

# Guarda el libro de trabajo de Excel en un archivo
    workbook.save('hotel.xlsx')
    workbook.close()'''