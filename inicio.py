import tkinter
import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import load_workbook
import win32com.client
from recursos import *

# Carga el archivo de Excel
workbook = load_workbook(filename='hotel.xlsx')

# Crea una nueva hoja en el archivo de Excel
sheet = workbook.create_sheet('Registro de clientes')

#excel = win32com.client.Dispatch('Excel.Application')
#instancia_excel = excel.Workbooks('hotel.xlsx')

window = tkinter.Tk()
window.title("Formulario para hotel")

frame = tkinter.Frame(window)
frame.pack()

user_info_frame =tkinter.LabelFrame(frame, text="Registrar cliente")
user_info_frame.grid(row= 0, column=0, padx=40, pady=15)

tipo_habitacion = tkinter.Label(user_info_frame, text="Tipo de Habitación")
tipo_habitacion.grid(row=0, column=0)

tipo_habitacion_combobox = ttk.Combobox(user_info_frame,values=['Simple','Doble','Triple','Matrimonial','Suite'])
tipo_habitacion_combobox.grid(row=1, column=0)

estado = tkinter.Label(user_info_frame, text="Estado")
estado.grid(row=0, column=1)

estado_combobox = ttk.Combobox(user_info_frame,values=['Desocupado/Limpieza','Habilitada','Reservado','Almacén','Pagado','Ocupada'])
estado_combobox.grid(row=1, column=1)

num_doc = tkinter.Label(user_info_frame, text="Nª Documento")
num_doc.grid(row=0, column=2)

num_doc_entrada = tkinter.Entry(user_info_frame)
num_doc_entrada.grid(row=1, column=2)

nombre = tkinter.Label(user_info_frame, text="Nombres")
nombre.grid(row=2, column=0)

nombre_entrada = tkinter.Entry(user_info_frame)
nombre_entrada.grid(row=3, column=0)

apellido = tkinter.Label(user_info_frame, text="Apellidos")
apellido.grid(row=2, column=1)

apellido_entrada = tkinter.Entry(user_info_frame)
apellido_entrada.grid(row=3, column=1)

num_boleta = tkinter.Label(user_info_frame, text="Nª Boleta")
num_boleta.grid(row=2, column=2)

num_boleta_entrada = tkinter.Entry(user_info_frame)
num_boleta_entrada.grid(row=3, column=2)

telefono = tkinter.Label(user_info_frame, text="Telefono")
telefono.grid(row=4, column=0)

telefono_entrada = tkinter.Entry(user_info_frame)
telefono_entrada.grid(row=5, column=0)

fecha_ingreso_entrada, fecha_salida_entrada,cond_indefinido,fecha_salida_indefinida = crear_widgets(user_info_frame,DateEntry)

user_info_frame_observacion =tkinter.LabelFrame(frame, text="Observación", padx=10, pady=10)
user_info_frame_observacion.grid(row= 6, column=0, padx=40, pady=15)

observacion_entrada = tkinter.Entry(user_info_frame_observacion, width=90)
observacion_entrada.grid(row=7, column=0, sticky="W")

for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=25, pady=5)

# Button
button = tkinter.Button(frame, text="Ingresar Datos", command=lambda: guardar_datos(tipo_habitacion_combobox,estado_combobox,num_doc_entrada,nombre_entrada,apellido_entrada,num_boleta_entrada,telefono_entrada,fecha_ingreso_entrada,fecha_salida_entrada,observacion_entrada,cond_indefinido,fecha_salida_indefinida))
button.grid(row=0, column=3, sticky="news", padx=20, pady=10)

window.mainloop()

#workbook.Close()
#excel.Quit()

