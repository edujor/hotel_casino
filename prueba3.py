import tkinter as tk
import openpyxl
from openpyxl import load_workbook

class ExcelEditor(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.grid()
        self.create_widgets()

    def create_widgets(self):
        # Crear las etiquetas para las columnas
        hab_label = tk.Label(self, text="HAB")
        hab_label.grid(row=0, column=0)
        nombre_label = tk.Label(self, text="NOMBRE")
        nombre_label.grid(row=0, column=1)
        dias_label = tk.Label(self, text="N° DIAS")
        dias_label.grid(row=0, column=2)

        # Crear la tabla utilizando una rejilla de entrada de datos (Entry)
        self.entries = []
        for i in range(5):  # Número de filas
            row_entries = []
            for j in range(3):  # Número de columnas
                entry = tk.Entry(self)
                entry.grid(row=i + 1, column=j)
                row_entries.append(entry)
            self.entries.append(row_entries)

        # Botón para cargar desde Excel
        load_button = tk.Button(self, text="Cargar Excel", command=self.load_from_excel)
        load_button.grid(row=6, column=0, columnspan=1)    

        # Botón para guardar en Excel
        save_button = tk.Button(self, text="Guardar", command=self.save_to_excel)
        save_button.grid(row=6, column=0, columnspan=3)

    def load_from_excel(self):
        workbook = load_workbook(filename='ejemplo3.xlsx')
        sheet = workbook.active

        # Leer los datos del archivo Excel y cargarlos en la tabla
        for i, row in enumerate(self.entries):
            for j, entry in enumerate(row):
                cell_value = sheet.cell(row=i + 2, column=j + 1).value
                entry.delete(0, tk.END)
                entry.insert(0, cell_value)

    def save_to_excel(self):
        # Crear un nuevo archivo de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Escribir los encabezados de las columnas en el archivo Excel
        headers = ["HAB", "NOMBRE", "N° DIAS"]
        for col, header in enumerate(headers):
            sheet.cell(row=1, column=col + 1).value = header

        # Recorrer las entradas de la tabla y guardar los valores en el archivo Excel
        for i, row in enumerate(self.entries):
            for j, entry in enumerate(row):
                value = entry.get()
                sheet.cell(row=i + 2, column=j + 1).value = value

        # Guardar el archivo Excel
        workbook.save('ejemplo3.xlsx')

# Crear la ventana principal
root = tk.Tk()
root.title("Editor de Excel")

# Crear la instancia de ExcelEditor
editor = ExcelEditor(root)
editor.pack()

# Iniciar el bucle principal de la ventana
root.mainloop()
