import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import load_workbook


class ExcelEditor(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.page_size = 10  # Cantidad de filas por página
        self.current_page = 0  # Página actual
        self.master = master
        self.grid()
        self.create_widgets()

    def create_widgets(self):
        # Crear las etiquetas para las columnas
        col1 = tk.Label(self, text="HAB")
        col1.grid(row=0, column=0)
        col2 = tk.Label(self, text="TIPO DE HAB")
        col2.grid(row=0, column=1)
        col3 = tk.Label(self, text="ESTADO")
        col3.grid(row=0, column=2)
        col4 = tk.Label(self, text="ID")
        col4.grid(row=0, column=3)
        col5 = tk.Label(self, text="APELLIDOS")
        col5.grid(row=0, column=4)
        col6 = tk.Label(self, text="NOMBRES")
        col6.grid(row=0, column=5)
        col7 = tk.Label(self, text="N° BOLETA")
        col7.grid(row=0, column=6)
        col8 = tk.Label(self, text="TELEFONO")
        col8.grid(row=0, column=7)
        col9 = tk.Label(self, text="F.INGRESO")
        col9.grid(row=0, column=8)
        col10 = tk.Label(self, text="F.SALIDA")
        col10.grid(row=0, column=9)


        # Crear la tabla utilizando una rejilla de entrada de datos (Entry)
        self.entries = []
        for i in range(20):  # Número de filas
            row_entries = []
            for j in range(10):  # Número de columnas
                entry = tk.Entry(self)
                entry.grid(row=i + 1, column=j)
                row_entries.append(entry)
            self.entries.append(row_entries)

        # Botón para cargar desde Excel
        load_button = tk.Button(self, text="Cargar Excel", command=self.load_from_excel)
        load_button.grid(row=6, column=0, columnspan=1)    

        # Botón para guardar en Excel
        save_button = tk.Button(self, text="Guardar", command=self.save_to_excel)
        save_button.grid(row=6, column=0, columnspan=3)

        # Botones de navegación para cargar las páginas
        next_button = tk.Button(self, text="Siguiente", command=self.next_page)
        next_button.grid(row=7, column=0)
        previous_button = tk.Button(self, text="Anterior", command=self.previous_page)
        previous_button.grid(row=7, column=1)

    def load_from_excel(self):
        # workbook = load_workbook(filename='ejemplo3.xlsx')
        workbook = load_workbook(filename='PRUEBA_1HOJA.xlsx')
        sheet = workbook.active
        # Calcular el rango de filas para la página actual
        start_row = self.current_page * self.page_size + 2  # +2 para omitir la primera fila de encabezados
        end_row = start_row + self.page_size

        # Leer los datos del archivo Excel y cargarlos en la tabla para la página actual
        for i, row in enumerate(self.entries):
            for j, entry in enumerate(row):
                cell_value = sheet.cell(row=i + start_row, column=j + 1).value
                if cell_value is not None:
                    entry.delete(0, tk.END)
                    entry.insert(0, cell_value)

        '''# Leer los datos del archivo Excel y cargarlos en la tabla
        for i, row in enumerate(self.entries):
            for j, entry in enumerate(row):
                cell_value = sheet.cell(row=i + 2, column=j + 1).value
                if cell_value is not None:
                    entry.delete(0, tk.END)
                    entry.insert(0, cell_value)'''
        
    def next_page(self):
        self.current_page += 1
        self.load_from_excel()

    def previous_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.load_from_excel()

    def save_to_excel(self):
        # Crear un nuevo archivo de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Escribir los encabezados de las columnas en el archivo Excel
        headers = ["HAB", "NOMBRE", "N° DIAS"]
        for col, header in enumerate(headers):
            sheet.cell(row=1, column=col + 1).value = header

        # Recorrer las entradas de la tabla y guardar los valores en el archivo Excel
        for i, row in enumerate(self.entries):
            for j, entry in enumerate(row):
                value = entry.get()
                sheet.cell(row=i + 2, column=j + 1).value = value

        # Guardar el archivo Excel
        workbook.save('ejemplo3.xlsx')

# Crear la ventana principal
root = tk.Tk()
root.title("Editor de Excel")

# Crear la instancia de ExcelEditor
editor = ExcelEditor(root)
editor.pack()

# Iniciar el bucle principal de la ventana
root.mainloop()